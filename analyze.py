import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import math
from collections import Counter

wb = openpyxl.load_workbook('yamap.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    r = dict(zip(headers, row))
    rows.append(r)

priority_cats = [
    'Красота', 'Строительство и недвижимость', 'Бизнес',
    'Услуги', 'Отдых и туризм', 'Медицина и фармацевтика',
    'Наука и образование', 'Культура и искусство', 'Семья'
]

low_cats = ['Коммунальные службы', 'Производство и поставки']

scored = []
for r in rows:
    rating = r.get('Рейтинг') or 0
    reviews = r.get('Кол-во отзывов') or 0
    scores_n = r.get('Кол-во оценок') or 0
    name = r.get('Название', '') or ''
    rub = str(r.get('Рубрика', '') or '')
    sub = str(r.get('Подрубрика', '') or '')

    try:
        rating = float(rating)
        reviews = int(reviews)
        scores_n = int(scores_n)
    except:
        rating = 0
        reviews = 0
        scores_n = 0

    score = rating * math.log(reviews + 1 + scores_n)

    for cat in priority_cats:
        if cat in rub:
            score *= 1.3
            break

    for cat in low_cats:
        if cat in rub or cat in sub:
            score *= 0.3
            break

    if any(x in name.lower() for x in ['жилищник', 'управляющая', 'гбу', 'муп', 'фгуп', 'мфц']):
        score *= 0.1

    scored.append((score, r))

scored.sort(key=lambda x: -x[0])

print('=== TOP 30 PROSPECTS ===')
for i, (score, r) in enumerate(scored[:30]):
    name = r['Название']
    rub = r['Рубрика']
    sub = r['Подрубрика']
    rat = r['Рейтинг']
    rev = r['Кол-во отзывов']
    oce = r['Кол-во оценок']
    tg = r['telegram']
    addr = r['Адрес']
    print(f'{i+1}. [{score:.1f}] {name}')
    print(f'   Рубрика: {rub}')
    print(f'   Подрубрика: {sub}')
    print(f'   Рейтинг: {rat} | Отзывов: {rev} | Оценок: {oce}')
    print(f'   Адрес: {addr}')
    print(f'   TG: {tg}')
    print()

# Also show subcategory breakdown for top prospects
print('=== SUBCATEGORY BREAKDOWN (top 100) ===')
sub_counter = Counter()
for score, r in scored[:100]:
    sub = str(r.get('Подрубрика', '') or '')
    for s in sub.split(','):
        sub_counter[s.strip()] += 1
for sub, count in sub_counter.most_common(20):
    print(f'  {sub}: {count}')
